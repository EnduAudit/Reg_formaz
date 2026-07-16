"""
Gestione Formazioni - registra ore di formazione su un file Excel condiviso
senza alterarne la formattazione (usa openpyxl, non riscrive lo stile).
"""
import json
import os
import sys
import datetime as dt
from copy import copy

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.comments import Comment

CONFIG_NAME = "gestione_formazioni_config.json"


def config_path():
    # accanto all'eseguibile/script, cosi' sopravvive tra un avvio e l'altro
    base = os.path.dirname(os.path.abspath(sys.argv[0]))
    return os.path.join(base, CONFIG_NAME)


def load_config():
    p = config_path()
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(cfg):
    with open(config_path(), "w", encoding="utf-8") as f:
        json.dump(cfg, f)


def find_header_layout(ws):
    """Individua colonne data (prima/ultima) e colonna TOTALE nella riga 1."""
    first_date_col = None
    last_date_col = None
    total_col = None
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, (dt.datetime, dt.date)):
            if first_date_col is None:
                first_date_col = c
            last_date_col = c
        elif isinstance(v, str) and v.strip().upper() == "TOTALE":
            total_col = c
    if first_date_col is None:
        raise ValueError("Non trovo colonne data nella riga 1 del file.")
    if total_col is None:
        total_col = last_date_col + 1
    return first_date_col, last_date_col, total_col


def build_date_map(ws, first_col, last_col):
    m = {}
    for c in range(first_col, last_col + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, (dt.datetime, dt.date)):
            iso = v.strftime("%Y-%m-%d") if isinstance(v, dt.datetime) else v.isoformat()
            m[iso] = c
    return m


def load_operators(ws):
    ops = []
    for r in range(2, ws.max_row + 1):
        nom = ws.cell(row=r, column=2).value
        if nom:
            ops.append((r, str(nom).strip()))
    return ops


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestione formazioni")
        self.geometry("560x640")
        self.resizable(False, False)

        self.file_path = None
        self.selected_row = None
        self.selected_name = None
        self.first_date_col = None
        self.last_date_col = None
        self.total_col = None
        self.session_log = []

        self._build_ui()
        self._try_autoload()

    # ---------- UI ----------
    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        top = tk.Frame(self)
        top.pack(fill="x", **pad)
        tk.Label(top, text="File Excel:", font=("Segoe UI", 9, "bold")).pack(side="left")
        self.file_label = tk.Label(top, text="(nessuno)", fg="#555")
        self.file_label.pack(side="left", padx=6)
        tk.Button(top, text="Seleziona file...", command=self.choose_file).pack(side="right")
        tk.Button(top, text="Ricarica", command=self.reload_file).pack(side="right", padx=6)

        sep = ttk.Separator(self)
        sep.pack(fill="x", padx=12, pady=4)

        # Ricerca operatore
        frm1 = tk.LabelFrame(self, text="1. Operatore", padx=10, pady=10)
        frm1.pack(fill="x", **pad)
        tk.Label(frm1, text="Cerca nominativo:").grid(row=0, column=0, sticky="w")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)
        search_entry = tk.Entry(frm1, textvariable=self.search_var, width=40)
        search_entry.grid(row=1, column=0, sticky="we", pady=2)

        self.listbox = tk.Listbox(frm1, height=5, width=55)
        self.listbox.grid(row=2, column=0, sticky="we", pady=2)
        self.listbox.bind("<<ListboxSelect>>", self._on_select_operator)

        self.selected_label = tk.Label(frm1, text="Nessun operatore selezionato", fg="#0f6e56", font=("Segoe UI", 9, "bold"))
        self.selected_label.grid(row=3, column=0, sticky="w", pady=(6, 0))

        tk.Button(frm1, text="+ Aggiungi nuovo nominativo", command=self.open_new_operator).grid(row=4, column=0, sticky="w", pady=(6, 0))

        # Nuovo operatore (nascosto di default)
        self.new_op_frame = tk.LabelFrame(self, text="Nuovo operatore", padx=10, pady=10)
        self.new_op_vars = {}
        fields = ["Nominativo", "Filiale", "Genere", "Inquadramento", "HC", "Funzione", "Mansione"]
        for i, f in enumerate(fields):
            tk.Label(self.new_op_frame, text=f + ":").grid(row=i, column=0, sticky="w")
            v = tk.StringVar()
            tk.Entry(self.new_op_frame, textvariable=v, width=35).grid(row=i, column=1, sticky="w", pady=1)
            self.new_op_vars[f] = v
        btns = tk.Frame(self.new_op_frame)
        btns.grid(row=len(fields), column=0, columnspan=2, pady=(8, 0))
        tk.Button(btns, text="Crea operatore", command=self.create_operator).pack(side="left", padx=4)
        tk.Button(btns, text="Annulla", command=self.close_new_operator).pack(side="left", padx=4)

        # Dettagli formazione
        frm2 = tk.LabelFrame(self, text="2. Dettagli formazione", padx=10, pady=10)
        frm2.pack(fill="x", **pad)

        tk.Label(frm2, text="Data (gg/mm/aaaa):").grid(row=0, column=0, sticky="w")
        self.date_var = tk.StringVar()
        tk.Entry(frm2, textvariable=self.date_var, width=15).grid(row=0, column=1, sticky="w")

        tk.Label(frm2, text="Ore:").grid(row=1, column=0, sticky="w")
        self.hours_var = tk.StringVar()
        tk.Entry(frm2, textvariable=self.hours_var, width=10).grid(row=1, column=1, sticky="w")

        tk.Label(frm2, text="Chi ha svolto la formazione:").grid(row=2, column=0, sticky="w")
        self.author_var = tk.StringVar()
        tk.Entry(frm2, textvariable=self.author_var, width=35).grid(row=2, column=1, sticky="w", columnspan=2)

        tk.Label(frm2, text="Argomento / postazione:").grid(row=3, column=0, sticky="w")
        self.topic_var = tk.StringVar()
        tk.Entry(frm2, textvariable=self.topic_var, width=35).grid(row=3, column=1, sticky="w", columnspan=2)

        tk.Button(frm2, text="Registra formazione", bg="#0f6e56", fg="white",
                  command=self.register_training).grid(row=4, column=0, columnspan=3, pady=(10, 0))

        # Log sessione
        frm3 = tk.LabelFrame(self, text="3. Formazioni registrate in questa sessione", padx=10, pady=10)
        frm3.pack(fill="both", expand=True, **pad)
        self.log_box = tk.Listbox(frm3, height=6)
        self.log_box.pack(fill="both", expand=True)

        self.status = tk.Label(self, text="", fg="#993c1d", wraplength=520, justify="left")
        self.status.pack(fill="x", padx=12, pady=(0, 10))

    # ---------- file handling ----------
    def _try_autoload(self):
        cfg = load_config()
        path = cfg.get("file_path")
        if path and os.path.exists(path):
            self.file_path = path
            self.file_label.config(text=os.path.basename(path))
            self.reload_file()

    def choose_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        self.file_path = path
        self.file_label.config(text=os.path.basename(path))
        save_config({"file_path": path})
        self.reload_file()

    def reload_file(self):
        if not self.file_path:
            messagebox.showwarning("Attenzione", "Seleziona prima un file.")
            return
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[wb.sheetnames[0]]
            self.first_date_col, self.last_date_col, self.total_col = find_header_layout(ws)
            self.date_map = build_date_map(ws, self.first_date_col, self.last_date_col)
            self.operators = load_operators(ws)
            self.status.config(text="")
            self.set_status_ok(
                f"File caricato: {len(self.operators)} operatori, periodo "
                f"{min(self.date_map)} - {max(self.date_map)}."
            )
        except PermissionError:
            self.status.config(text="Il file è aperto/bloccato da un altro programma. Chiudilo e riprova.")
        except Exception as e:
            self.status.config(text=f"Errore nel caricamento: {e}")

    def set_status_ok(self, text):
        self.status.config(text=text, fg="#0f6e56")

    def set_status_err(self, text):
        self.status.config(text=text, fg="#993c1d")

    # ---------- ricerca operatore ----------
    def _on_search(self, *a):
        q = self.search_var.get().strip().lower()
        self.listbox.delete(0, tk.END)
        if not q or not hasattr(self, "operators"):
            return
        matches = [o for o in self.operators if q in o[1].lower()][:15]
        self._current_matches = matches
        for _, name in matches:
            self.listbox.insert(tk.END, name)

    def _on_select_operator(self, event):
        sel = self.listbox.curselection()
        if not sel:
            return
        row, name = self._current_matches[sel[0]]
        self.selected_row = row
        self.selected_name = name
        self.selected_label.config(text=f"Selezionato: {name}")

    # ---------- nuovo operatore ----------
    def open_new_operator(self):
        self.new_op_vars["Nominativo"].set(self.search_var.get().strip())
        self.new_op_frame.pack(fill="x", padx=12, pady=6, before=self.winfo_children()[3])

    def close_new_operator(self):
        self.new_op_frame.pack_forget()

    def create_operator(self):
        name = self.new_op_vars["Nominativo"].get().strip()
        if not name:
            messagebox.showwarning("Attenzione", "Inserisci il nominativo.")
            return
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[wb.sheetnames[0]]
            new_row = ws.max_row + 1
            template_row = ws.max_row  # ultima riga esistente, per copiare lo stile

            values = {
                1: self.new_op_vars["Filiale"].get().strip(),
                2: name,
                3: self.new_op_vars["Genere"].get().strip(),
                4: self.new_op_vars["Inquadramento"].get().strip(),
                5: self.new_op_vars["HC"].get().strip(),
                6: self.new_op_vars["Funzione"].get().strip(),
                7: self.new_op_vars["Mansione"].get().strip(),
            }
            for col in range(1, self.total_col + 1):
                src = ws.cell(row=template_row, column=col)
                dst = ws.cell(row=new_row, column=col)
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
            for col, val in values.items():
                if val:
                    ws.cell(row=new_row, column=col).value = val

            first_letter = openpyxl.utils.get_column_letter(self.first_date_col)
            last_letter = openpyxl.utils.get_column_letter(self.last_date_col)
            ws.cell(row=new_row, column=self.total_col).value = (
                f"=+SUM({first_letter}{new_row}:{last_letter}{new_row})"
            )
            ws.row_dimensions[new_row].height = ws.row_dimensions[template_row].height

            wb.save(self.file_path)
            self.operators.append((new_row, name))
            self.close_new_operator()
            for v in self.new_op_vars.values():
                v.set("")
            self.selected_row = new_row
            self.selected_name = name
            self.selected_label.config(text=f"Selezionato: {name}")
            self.set_status_ok(f"Operatore '{name}' creato e salvato nel file.")
        except PermissionError:
            self.set_status_err("Il file è aperto/bloccato da un altro programma. Chiudilo e riprova.")
        except Exception as e:
            self.set_status_err(f"Errore: {e}")

    # ---------- registrazione formazione ----------
    def register_training(self):
        if self.selected_row is None:
            messagebox.showwarning("Attenzione", "Seleziona prima un operatore.")
            return
        date_str = self.date_var.get().strip()
        try:
            d = dt.datetime.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Data non valida", "Usa il formato gg/mm/aaaa.")
            return
        try:
            hours = float(self.hours_var.get().replace(",", "."))
            if hours <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ore non valide", "Inserisci un numero di ore valido.")
            return
        author = self.author_var.get().strip()
        topic = self.topic_var.get().strip()
        if not author or not topic:
            messagebox.showwarning("Attenzione", "Indica chi ha svolto la formazione e l'argomento.")
            return

        iso = d.strftime("%Y-%m-%d")

        try:
            # ricarico il file appena prima di scrivere, per ridurre conflitti
            # con altri operatori che potrebbero averlo modificato nel frattempo
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[wb.sheetnames[0]]
            first_date_col, last_date_col, total_col = find_header_layout(ws)
            date_map = build_date_map(ws, first_date_col, last_date_col)
            if iso not in date_map:
                self.set_status_err(f"La data {date_str} non è presente nel file.")
                return
            col = date_map[iso]

            # ritrovo la riga dell'operatore per nome, nel caso siano cambiate le righe
            row = None
            for r in range(2, ws.max_row + 1):
                nom = ws.cell(row=r, column=2).value
                if nom and str(nom).strip() == self.selected_name:
                    row = r
                    break
            if row is None:
                self.set_status_err("Operatore non più trovato nel file (forse è stato rimosso). Ricarica.")
                return

            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                ok = messagebox.askyesno(
                    "Cella già compilata",
                    f"{self.selected_name} in data {date_str} ha già {cell.value} ore.\n"
                    f"Sostituire con {hours} ore?"
                )
                if not ok:
                    return

            previous = cell.value
            cell.value = hours
            cell.comment = Comment(f"{author}: {topic}", author)

            wb.save(self.file_path)

            self.session_log.append((self.selected_name, date_str, hours, author, topic))
            self.log_box.insert(
                tk.END, f"{self.selected_name} - {date_str} - {hours}h - {author} - {topic}"
            )
            self.set_status_ok("Formazione registrata e salvata nel file.")
            self.hours_var.set("")
            self.author_var.set("")
            self.topic_var.set("")
        except PermissionError:
            self.set_status_err(
                "Il file è aperto/bloccato (es. qualcuno lo ha aperto in Excel, oppure un altro "
                "operatore sta salvando in questo momento). Riprova tra qualche secondo."
            )
        except Exception as e:
            self.set_status_err(f"Errore durante il salvataggio: {e}")


if __name__ == "__main__":
    App().mainloop()
